# Using an excel file to do random stuff
from tkinter import *
from tkinter.ttk import *
from openpyxl import workbook, load_workbook
import random

root = Tk()  # Tk allows us to create a basic pop up window
root.title("Random Name Lottery")
root.geometry("400x400")  # how big is the window
root.resizable(0, 0)  # can't expand the window


def randomname(event):
    wb = load_workbook("hello.xslx")  # make sure file is in the folder first!
    ws = wb.active  # worksheet
    rangeline = ws["A2":"A19"]  # creates a list
    name = []  # creating an empy variable ready to take LIST
    for items in rangeline:  # gives cell value
        for subitems in items:  # gets what's inside the cell
            name.append(subitems.value)  # change and save to NAME variable
            print(name)
    computer_action = random.choice(name)  # add random function
    print("the random person is: " + name)

    updateDisplay(computer_action)


def updateDisplay(abc):
    displayVariable.set(abc)


button_1 = Button(root, text="Random Lottery Name")
button_1.bind("<button-1>", randomname)
button_1.pack()
displayVariable = StringVar()
displayLabel = Label(root, textvariable=displayVariable)
displayLabel.pack()

mainloop()
